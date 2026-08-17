import asyncio
import csv
import re
from pathlib import Path
from uuid import uuid4

from aiogram import Bot, F, Router
from aiogram.types import FSInputFile, Message
from loguru import logger

from app.bot.filters import ShouldRespondFilter
from app.bot.handlers.admin import ACCESS_DENIED_MESSAGE, is_admin
from app.core.config import settings
from app.core.database import async_session_maker
from app.core.router import CascadeRouter
from app.models.sqlalchemy.document import Document
from app.models.sqlalchemy.engineering_doc import EngineeringDoc
from app.models.sqlalchemy.project import Project
from app.models.sqlalchemy.project_file import ProjectFile
from app.models.sqlalchemy.user import User
from app.services import cad_parser
from app.services.engineering_rag_ingest import index_engineering_doc
from app.services.office_parser import extract_docx_text, extract_pptx_text
from app.services.pdf_parser import chunk_text, extract_text
from app.services.stock_import import StockTableError, parse_stock_table, upsert_stock_rows

router = Router(name="documents")

DOCUMENT_TEMP_DIR = Path("data/temp")
CAD_EXTENSIONS = {".dxf", ".dwg", ".cdr"}
STOCK_TABLE_EXTENSIONS = {".xlsx", ".csv"}
TEXT_EXTENSIONS = {".txt", ".md"}
OFFICE_EXTRACTORS = {".docx": extract_docx_text, ".pptx": extract_pptx_text}

# Captioning an upload with "проект3 <ID>" attaches it to that Project as a
# config/preset file (ProjectFile) regardless of extension — the one place
# arbitrary non-CAD project files actually get in, since Telegram commands
# can't carry a file attachment themselves. Explicit caption intent wins
# over extension-based routing (checked first in handle_document).
PROJECT_FILE_CAPTION_PATTERN = re.compile(r"\bпроект3\s+(\d+)\b", re.IGNORECASE)


def _is_pdf(file_name: str | None, mime_type: str | None) -> bool:
    if mime_type == "application/pdf":
        return True
    return bool(file_name) and file_name.lower().endswith(".pdf")


def _cad_extension(file_name: str | None) -> str | None:
    if not file_name:
        return None
    ext = Path(file_name).suffix.lower()
    return ext if ext in CAD_EXTENSIONS else None


def _stock_table_extension(file_name: str | None) -> str | None:
    if not file_name:
        return None
    ext = Path(file_name).suffix.lower()
    return ext if ext in STOCK_TABLE_EXTENSIONS else None


def _text_extension(file_name: str | None) -> str | None:
    if not file_name:
        return None
    ext = Path(file_name).suffix.lower()
    return ext if ext in TEXT_EXTENSIONS else None


def _office_extension(file_name: str | None) -> str | None:
    if not file_name:
        return None
    ext = Path(file_name).suffix.lower()
    return ext if ext in OFFICE_EXTRACTORS else None


@router.message(F.document, ShouldRespondFilter())
async def handle_document(message: Message, bot: Bot, cascade_router: CascadeRouter, db_user: User) -> None:
    document = message.document

    caption_match = PROJECT_FILE_CAPTION_PATTERN.search(message.caption or "")
    if caption_match:
        await _handle_project_file_upload(message, bot, int(caption_match.group(1)))
        return

    if _is_pdf(document.file_name, document.mime_type):
        await _handle_pdf_upload(message, bot, cascade_router, db_user)
        return

    cad_ext = _cad_extension(document.file_name)
    if cad_ext:
        await _handle_cad_upload(message, bot, cad_ext, cascade_router)
        return

    stock_ext = _stock_table_extension(document.file_name)
    if stock_ext:
        await _handle_stock_table_upload(message, bot, stock_ext)
        return

    text_ext = _text_extension(document.file_name)
    if text_ext:
        await _handle_text_upload(message, bot, cascade_router, db_user, text_ext)
        return

    office_ext = _office_extension(document.file_name)
    if office_ext:
        await _handle_office_upload(message, bot, cascade_router, db_user, office_ext)
        return

    await message.answer(
        "Пока поддерживается загрузка PDF, .txt/.md, .docx, .pptx, .dxf, .dwg, .xlsx/.csv (остатки склада) — "
        ".cdr не читается ни одним инструментом."
    )


async def _handle_pdf_upload(message: Message, bot: Bot, cascade_router: CascadeRouter, db_user: User) -> None:
    document = message.document
    file = await bot.get_file(document.file_id)
    DOCUMENT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    local_path = DOCUMENT_TEMP_DIR / f"doc_{uuid4().hex}.pdf"

    try:
        await bot.download_file(file.file_path, destination=local_path)
        text = await asyncio.to_thread(extract_text, str(local_path))
    finally:
        local_path.unlink(missing_ok=True)

    chunks = chunk_text(text)
    if not chunks:
        await message.answer("Не удалось извлечь текст из документа.")
        return

    filename = document.file_name or "document.pdf"
    cascade_router.rag_engine.add_documents(
        texts=chunks,
        metadatas=[{"source": "pdf_upload", "filename": filename, "uploaded_by": str(db_user.id)} for _ in chunks],
    )

    async with async_session_maker() as session:
        session.add(
            Document(
                source="pdf_upload",
                filename=filename,
                uploaded_by=db_user.id,
                chunk_count=len(chunks),
                char_count=len(text),
                embedding_model=cascade_router.rag_engine.embedding_model_name,
            )
        )
        await session.commit()

    await message.answer(f"Документ «{filename}» обработан и добавлен в базу знаний ({len(chunks)} фрагм.).")


def _decode_text_file(raw: bytes) -> str:
    # Plain .txt/.md files from Russian-speaking users are as likely to be
    # cp1251 as UTF-8 (Notepad's historical default) — try both before
    # falling back to a lossy decode rather than erroring on the whole
    # upload over one file's encoding.
    for encoding in ("utf-8", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


async def _handle_text_upload(
    message: Message, bot: Bot, cascade_router: CascadeRouter, db_user: User, ext: str
) -> None:
    document = message.document
    file = await bot.get_file(document.file_id)
    DOCUMENT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    local_path = DOCUMENT_TEMP_DIR / f"text_{uuid4().hex}{ext}"

    try:
        await bot.download_file(file.file_path, destination=local_path)
        text = await asyncio.to_thread(lambda: _decode_text_file(local_path.read_bytes()))
    finally:
        local_path.unlink(missing_ok=True)

    chunks = chunk_text(text)
    if not chunks:
        await message.answer("Файл пустой — нечего добавлять в базу знаний.")
        return

    filename = document.file_name or f"document{ext}"
    cascade_router.rag_engine.add_documents(
        texts=chunks,
        metadatas=[{"source": "text_upload", "filename": filename, "uploaded_by": str(db_user.id)} for _ in chunks],
    )

    async with async_session_maker() as session:
        session.add(
            Document(
                source="text_upload",
                filename=filename,
                uploaded_by=db_user.id,
                chunk_count=len(chunks),
                char_count=len(text),
                embedding_model=cascade_router.rag_engine.embedding_model_name,
            )
        )
        await session.commit()

    await message.answer(f"Документ «{filename}» обработан и добавлен в базу знаний ({len(chunks)} фрагм.).")


async def _handle_office_upload(
    message: Message, bot: Bot, cascade_router: CascadeRouter, db_user: User, ext: str
) -> None:
    document = message.document
    file = await bot.get_file(document.file_id)
    DOCUMENT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    local_path = DOCUMENT_TEMP_DIR / f"office_{uuid4().hex}{ext}"

    try:
        await bot.download_file(file.file_path, destination=local_path)
        text = await asyncio.to_thread(OFFICE_EXTRACTORS[ext], str(local_path))
    finally:
        local_path.unlink(missing_ok=True)

    chunks = chunk_text(text)
    if not chunks:
        await message.answer("Не удалось извлечь текст из документа.")
        return

    source = "docx_upload" if ext == ".docx" else "pptx_upload"
    filename = document.file_name or f"document{ext}"
    cascade_router.rag_engine.add_documents(
        texts=chunks,
        metadatas=[{"source": source, "filename": filename, "uploaded_by": str(db_user.id)} for _ in chunks],
    )

    async with async_session_maker() as session:
        session.add(
            Document(
                source=source,
                filename=filename,
                uploaded_by=db_user.id,
                chunk_count=len(chunks),
                char_count=len(text),
                embedding_model=cascade_router.rag_engine.embedding_model_name,
            )
        )
        await session.commit()

    await message.answer(f"Документ «{filename}» обработан и добавлен в базу знаний ({len(chunks)} фрагм.).")


def _format_cad_summary(project_name: str, extracted: cad_parser.ExtractedCadData) -> str:
    lines = [f"Чертёж «{project_name}» разобран."]
    if extracted.entity_counts:
        counts = ", ".join(f"{name}: {count}" for name, count in extracted.entity_counts.items())
        lines.append(f"Элементы: {counts}")
    if extracted.dimensions:
        lines.append(f"Размеры: {', '.join(extracted.dimensions[:10])}")
    if extracted.texts:
        lines.append(f"Текст на чертеже: {', '.join(extracted.texts[:10])}")
    return "\n".join(lines)


async def _handle_cad_upload(message: Message, bot: Bot, ext: str, cascade_router: CascadeRouter) -> None:
    document = message.document
    file = await bot.get_file(document.file_id)
    DOCUMENT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    local_path = DOCUMENT_TEMP_DIR / f"cad_{uuid4().hex}{ext}"

    try:
        await bot.download_file(file.file_path, destination=local_path)
    except Exception:
        local_path.unlink(missing_ok=True)
        raise

    project_name = Path(document.file_name).stem
    await message.answer(f"⏳ Чертёж «{project_name}» принят, разбираю и рендерю...")

    # Backgrounded (see media.py's _run_download for the same
    # asyncio.create_task pattern) so parsing + rendering a large/complex
    # drawing never risks the aiogram update-processing timeout, and the
    # user gets an immediate ack instead of a silently stalled chat while
    # ezdxf/matplotlib do their (potentially slow) work off-thread.
    asyncio.create_task(_process_cad_upload(message, local_path, project_name, cascade_router))


async def _process_cad_upload(
    message: Message, local_path: Path, project_name: str, cascade_router: CascadeRouter
) -> None:
    # Runs detached via asyncio.create_task — nothing awaits this coroutine,
    # so an exception that escapes it doesn't propagate anywhere a human
    # would see it (asyncio just logs "Task exception was never retrieved"
    # on GC). Every path below must therefore end in a reply to the user,
    # success or failure, rather than letting anything fall through.
    try:
        try:
            doc, doc_type = await asyncio.to_thread(
                cad_parser.open_drawing, local_path, settings.ODA_FILE_CONVERTER_PATH
            )
        except (cad_parser.UnsupportedCadFormatError, cad_parser.CadConversionError, cad_parser.CadParseError) as exc:
            logger.warning(f"CAD upload failed for {project_name}: {exc}")
            await message.answer(str(exc))
            return

        extracted = await asyncio.to_thread(cad_parser.extract_data, doc)

        cad_storage = Path(settings.CAD_STORAGE_PATH)
        cad_storage.mkdir(parents=True, exist_ok=True)
        stored_dxf = cad_storage / f"{uuid4().hex}.dxf"
        await asyncio.to_thread(doc.saveas, str(stored_dxf))
        rendered_pdf = stored_dxf.with_suffix(".pdf")
        await asyncio.to_thread(cad_parser.render_to_pdf, doc, rendered_pdf)

        async with async_session_maker() as session:
            engineering_doc = EngineeringDoc(
                project_name=project_name,
                file_path=str(stored_dxf),
                doc_type=doc_type,
                extracted_data=extracted.to_dict(),
                is_generated=False,
            )
            session.add(engineering_doc)
            await session.commit()
            await session.refresh(engineering_doc)
            index_engineering_doc(cascade_router.rag_engine, engineering_doc)

        await message.answer(_format_cad_summary(project_name, extracted))
        await message.answer_document(FSInputFile(str(rendered_pdf)))
    except Exception as exc:
        logger.exception(f"CAD upload background processing failed for {project_name}: {exc}")
        await message.answer(
            f"Не удалось обработать чертёж «{project_name}» — внутренняя ошибка. Попробуйте другой файл "
            "или обратитесь к администратору."
        )
    finally:
        local_path.unlink(missing_ok=True)


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    # Imported lazily — openpyxl is only needed for this one upload path.
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    return [["" if cell is None else str(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]


def _read_csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


async def _handle_project_file_upload(message: Message, bot: Bot, project_id: int) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(ACCESS_DENIED_MESSAGE)
        return

    async with async_session_maker() as session:
        project = await session.get(Project, project_id)
    if project is None:
        await message.answer(f"Проект #{project_id} не найден.")
        return

    document = message.document
    file = await bot.get_file(document.file_id)
    storage_dir = Path(settings.PROJECT_FILES_PATH)
    storage_dir.mkdir(parents=True, exist_ok=True)

    file_name = document.file_name or f"file_{uuid4().hex}"
    stored_path = storage_dir / f"{uuid4().hex}_{file_name}"
    try:
        await bot.download_file(file.file_path, destination=stored_path)
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise

    async with async_session_maker() as session:
        session.add(
            ProjectFile(project_id=project.id, file_path=str(stored_path), file_name=file_name, kind="config")
        )
        await session.commit()

    await message.answer(f"Файл «{file_name}» прикреплён к проекту «{project.name}».")


async def _handle_stock_table_upload(message: Message, bot: Bot, ext: str) -> None:
    if not is_admin(message.from_user.id):
        await message.answer(ACCESS_DENIED_MESSAGE)
        return

    document = message.document
    file = await bot.get_file(document.file_id)
    DOCUMENT_TEMP_DIR.mkdir(parents=True, exist_ok=True)
    local_path = DOCUMENT_TEMP_DIR / f"stock_{uuid4().hex}{ext}"

    try:
        await bot.download_file(file.file_path, destination=local_path)
        reader = _read_xlsx_rows if ext == ".xlsx" else _read_csv_rows
        rows = await asyncio.to_thread(reader, local_path)
    finally:
        local_path.unlink(missing_ok=True)

    try:
        stock_rows = parse_stock_table(rows)
    except StockTableError as exc:
        await message.answer(str(exc))
        return

    async with async_session_maker() as session:
        count = await upsert_stock_rows(session, stock_rows)

    await message.answer(f"Импортировано позиций: {count}.")
